"""
Mostly made by Claude

Reads a Tecan plate-reader export (.xlsx) with two sheets:
  - Sheet1: legend mapping a well-column NUMBER -> treatment label
            e.g.  1 | Mock
                  2 | 100nM flg22
                  3 | - luminol
  - Sheet2: raw kinetic luminescence data exported directly from the
            Tecan i-control software 

The well ROW letter (A, B, C ...) is treated as a biological/technical
replicate, and the well COLUMN number is matched to the legend to
determine treatment identity. Replicates are averaged at each timepoint
and plotted as a mean line +/- SEM ribbon.
"""

import re
import sys
import argparse
import numpy as np
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt



def load_legend(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    legend = {}
    for row in ws.iter_rows(values_only=True):
        if row[0] is None or row[1] is None:
            continue
        try:
            key = int(row[0])
        except (TypeError, ValueError):
            continue
        legend[key] = str(row[1]).strip()
    if not legend:
        raise ValueError("Could not parse a legend from Sheet1. Expected "
                          "two columns: well-column-number, treatment label.")
    return legend


def load_plate_data(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet2"]

    n_rows, n_cols = ws.max_row, ws.max_column
    col_a = [ws.cell(row=r, column=1).value for r in range(1, n_rows + 1)]

    # Locate the header rows by label in column A
    time_row = None
    for r, v in enumerate(col_a, start=1):
        if isinstance(v, str) and v.strip().lower().startswith("time ["):
            time_row = r
            break
    if time_row is None:
        raise ValueError("Could not find the 'Time [s]' header row in Sheet2.")

    # Time values run across the row, starting at column 2
    time_vals = []
    for c in range(2, n_cols + 1):
        v = ws.cell(row=time_row, column=c).value
        if v is None:
            break
        time_vals.append(float(v))
    time_vals = np.array(time_vals)
    time_minutes = time_vals / 60.0
    n_points = len(time_vals)

    # Well rows: label in column A that matches a well pattern like "A1", "H12"
    well_pat = re.compile(r"^[A-H](\d{1,2})$")
    wells = {}
    for r in range(time_row + 1, n_rows + 1):
        label = col_a[r - 1]
        if not isinstance(label, str):
            continue
        label = label.strip()
        if not well_pat.match(label):
            continue
        vals = [ws.cell(row=r, column=c).value for c in range(2, 2 + n_points)]
        if all(v is None for v in vals):
            continue
        vals = np.array([np.nan if v is None else float(v) for v in vals])
        wells[label] = vals

    if not wells:
        raise ValueError("Could not parse any well rows (e.g. 'A1', 'B3') in Sheet2.")

    return time_minutes, wells


def build_condition_table(time_minutes, wells, legend):
    """Group replicate wells by their column number -> legend condition.

    Returns a dict: condition_label -> DataFrame (replicates x timepoints)
    """
    well_pat = re.compile(r"^[A-H](\d{1,2})$")
    grouped = {}
    for well_id, vals in wells.items():
        m = well_pat.match(well_id)
        col_num = int(m.group(1))
        if col_num not in legend:
            continue  # column not present in legend -> not part of this experiment
        label = legend[col_num]
        grouped.setdefault(label, []).append(vals)

    tables = {}
    for label, rep_list in grouped.items():
        arr = np.vstack(rep_list)
        tables[label] = pd.DataFrame(arr, columns=time_minutes)
    return tables


def plot_ribbon(time_minutes, tables, legend, out_path, error="sem"):

    import matplotlib.font_manager as fm
    preferred_fonts = ["Arial", "Helvetica", "DejaVu Sans"]
    available = {f.name for f in fm.fontManager.ttflist}
    font_family = next((f for f in preferred_fonts if f in available), "sans-serif")

    plt.rcParams.update({
        "font.family": font_family,
        "font.size": 9,
        "axes.linewidth": 0.8,
        "xtick.major.width": 0.8,
        "ytick.major.width": 0.8,
        "pdf.fonttype": 42,
        "ps.fonttype": 42,
        "svg.fonttype": "none",
    })

    ordered_labels = [legend[k] for k in sorted(legend) if legend[k] in tables]

    accent_colors = ["#2A6FBB", "#C1443C", "#3E9B4F", "#8A5AAB", "#D89A2A"]

    mock_pat = re.compile(r"\bmock\b|\buntreated\b", re.IGNORECASE)

    color_map = {}
    accent_i = 0
    for label in ordered_labels:
        if mock_pat.search(label):
            color_map[label] = "#8C8C8C"  # grey
        else:
            color_map[label] = accent_colors[accent_i % len(accent_colors)]
            accent_i += 1

    fig, ax = plt.subplots(figsize=(3.6, 2.7), dpi=300)

    for label in ordered_labels:
        df = tables[label]
        mean = df.mean(axis=0).values
        n = df.shape[0]
        sd = df.std(axis=0, ddof=1).values
        spread = sd / np.sqrt(n) if error == "sem" else sd

        color = color_map[label]
        ax.plot(time_minutes, mean, color=color, linewidth=1.4,
                label=f"{label} (n={n})", zorder=3)
        ax.fill_between(time_minutes, mean - spread, mean + spread,
                         color=color, alpha=0.18, linewidth=0, zorder=2)

    ax.set_xlabel("Time (min)", labelpad=4)
    ax.set_ylabel("RLU", labelpad=4)
    ax.set_xlim(time_minutes.min(), time_minutes.max())

    # Nature-style minimal axes: no top/right spines, outward ticks
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.tick_params(direction="out", length=3)
    ax.margins(y=0.05)

    leg = ax.legend(frameon=False, fontsize=7, loc="upper left",
                     handlelength=1.4, borderaxespad=0)

    fig.tight_layout()
    fig.savefig(out_path, bbox_inches="tight")
    print(f"Saved figure to: {out_path}")

    # Also save a PNG alongside if the requested output is a PDF (or vice versa)
    alt_path = out_path.rsplit(".", 1)[0] + (".png" if out_path.endswith(".pdf") else ".pdf")
    fig.savefig(alt_path, bbox_inches="tight")
    print(f"Saved figure to: {alt_path}")

    plt.close(fig)


def main():
    parser = argparse.ArgumentParser(description="Plot ROS burst curve as a ribbon plot.")
    parser.add_argument("filepath", nargs="?", default=FILEPATH,
                         help="Path to the .xlsx file (defaults to FILEPATH set in the script).")
    parser.add_argument("--out", default=None,
                         help="Output file path (default: <input_name>_ROS_curve.pdf)")
    parser.add_argument("--sem", choices=["sem", "sd"], default="sem",
                         help="Error band type: 'sem' (default) or 'sd'.")
    args = parser.parse_args()

    legend = load_legend(args.filepath)
    print("Legend loaded from Sheet1:")
    for k, v in sorted(legend.items()):
        print(f"  {k} -> {v}")

    time_minutes, wells = load_plate_data(args.filepath)
    print(f"Parsed {len(wells)} wells across {len(time_minutes)} timepoints "
          f"({time_minutes[-1]:.1f} min total).")

    tables = build_condition_table(time_minutes, wells, legend)
    print("Conditions found:", {k: v.shape[0] for k, v in tables.items()}, "replicates")

    out_path = args.out
    if out_path is None:
        base = args.filepath.rsplit("/", 1)[-1].rsplit(".", 1)[0]
        out_path = f"{base}_ROS_curve.pdf"

    plot_ribbon(time_minutes, tables, legend, out_path, error=args.sem)


if __name__ == "__main__":
    main()